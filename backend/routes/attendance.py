from flask import Blueprint, jsonify, send_file
from database import get_all_attendance, get_all_students
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
from datetime import datetime

attendance_bp = Blueprint("attendance", __name__)

@attendance_bp.route("/attendance", methods=["GET"])
def attendance():
    try:
        records = get_all_attendance()
        attendance_list = []
        for record in records:
            attendance_list.append({
                "id": record[0],
                "name": record[1],
                "date": record[2],
                "time": record[3]
            })
        return jsonify({
            "status": "success",
            "total": len(attendance_list),
            "attendance": attendance_list
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@attendance_bp.route("/export", methods=["GET"])
def export_attendance():
    try:
        records = get_all_attendance()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")

        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = "Attendance Report"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:D2")
        date_cell = ws["A2"]
        date_cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.alignment = Alignment(horizontal="center")

        ws.append([])
        headers = ["#", "Student Name", "Date", "Time"]
        ws.append(headers)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, record in enumerate(records, 1):
            ws.append([i, record[1], record[2], record[3]])
            fill_color = "E3F2FD" if i % 2 == 0 else "FFFFFF"
            for col in range(1, 5):
                cell = ws.cell(row=4 + i, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"attendance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@attendance_bp.route("/dashboard", methods=["GET"])
def dashboard():
    try:
        records = get_all_attendance()
        today = datetime.now().strftime("%Y-%m-%d")

        total_records = len(records)
        today_records = [r for r in records if r[2] == today]
        total_today = len(today_records)
        all_students = list(set([r[1] for r in records]))
        total_students = len(all_students)

        student_counts = {}
        for record in records:
            name = record[1]
            student_counts[name] = student_counts.get(name, 0) + 1

        today_names = [r[1] for r in today_records]

        return jsonify({
            "status": "success",
            "total_records": total_records,
            "total_students": total_students,
            "total_today": total_today,
            "today_date": today,
            "today_names": today_names,
            "student_counts": student_counts
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@attendance_bp.route("/absent", methods=["GET"])
def get_absent():
    try:
        today = datetime.now().strftime("%Y-%m-%d")
        all_students = get_all_students()
        records = get_all_attendance()
        present_today = list(set([r[1] for r in records if r[2] == today]))
        absent_today = [s for s in all_students if s not in present_today]

        return jsonify({
            "status": "success",
            "today": today,
            "present": present_today,
            "absent": absent_today,
            "total_students": len(all_students),
            "total_present": len(present_today),
            "total_absent": len(absent_today)
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500